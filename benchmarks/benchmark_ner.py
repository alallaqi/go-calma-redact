"""
NER Model Benchmark
===================
Standalone script — does NOT modify the main application.

Usage:
    python benchmarks/benchmark_ner.py --docs path/to/pdf/folder [--out benchmarks/results]

For each available NER model it will:
  1. Pre-extract text from all PDFs once (text extraction time excluded from NER timing)
  2. Run detection on every document and measure wall-clock time
  3. Count total entities and break them down by entity type
  4. Write results to <out>.json and <out>.xlsx

Requirements (in addition to the main project deps):
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Ensure the project root is on sys.path so `gocalma` can be imported
# when the script is run from any working directory.
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from gocalma.pdf_extract import extract_text
from gocalma.pii_detect import available_models, detect_pii_all_pages, NLP_MODELS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_pdfs(folder: Path, limit: int = 30) -> list[dict]:
    """Return a list of dicts with keys: name, path, pages (extracted text)."""
    pdf_paths = sorted(folder.glob("**/*.pdf"))[:limit]
    if not pdf_paths:
        print(f"[!] No PDFs found in {folder}")
        sys.exit(1)

    docs = []
    print(f"[+] Pre-extracting text from {len(pdf_paths)} PDF(s)...")
    for p in pdf_paths:
        raw = p.read_bytes()
        pages = extract_text(raw)
        docs.append({"name": p.name, "path": str(p), "pages": pages})
        print(f"    {p.name}  ({len(pages)} page(s))")
    return docs


def run_model(model_key: str, docs: list[dict], score_threshold: float) -> list[dict]:
    """Run *model_key* over every doc and return per-doc result rows."""
    rows = []
    lang = NLP_MODELS[model_key]["lang_codes"][0]

    for i, doc in enumerate(docs):
        t0 = time.perf_counter()
        entities = detect_pii_all_pages(
            doc["pages"],
            language=lang,
            score_threshold=score_threshold,
            model_key=model_key,
        )
        elapsed = time.perf_counter() - t0

        # Entity type breakdown
        type_counts: dict[str, int] = {}
        for e in entities:
            type_counts[e.entity_type] = type_counts.get(e.entity_type, 0) + 1

        rows.append({
            "model": model_key,
            "document": doc["name"],
            "pages": len(doc["pages"]),
            "time_s": round(elapsed, 3),
            "time_per_page_s": round(elapsed / max(len(doc["pages"]), 1), 3),
            "entities_total": len(entities),
            "entity_types": type_counts,
            # First doc includes model cold-load time; flag it for transparency
            "includes_model_load": i == 0,
        })

        status = "(incl. model load)" if i == 0 else ""
        print(
            f"    [{i+1}/{len(docs)}] {doc['name']}: "
            f"{len(entities)} entities in {elapsed:.2f}s {status}"
        )

    return rows


def summarise(model_key: str, rows: list[dict]) -> dict:
    """Compute aggregate stats for a model from its per-doc rows."""
    times = [r["time_s"] for r in rows]
    # Exclude first doc from steady-state averages to avoid model-load bias
    steady = [r["time_s"] for r in rows[1:]] if len(rows) > 1 else times
    all_entity_counts = [r["entities_total"] for r in rows]
    all_types: dict[str, int] = {}
    for r in rows:
        for t, c in r["entity_types"].items():
            all_types[t] = all_types.get(t, 0) + c

    return {
        "model": model_key,
        "docs_processed": len(rows),
        "total_entities": sum(all_entity_counts),
        "avg_entities_per_doc": round(sum(all_entity_counts) / len(rows), 1),
        "avg_time_per_doc_s": round(sum(times) / len(times), 3),
        "avg_time_per_doc_steady_s": round(sum(steady) / len(steady), 3) if steady else None,
        "min_time_s": round(min(times), 3),
        "max_time_s": round(max(times), 3),
        "cold_load_time_s": round(rows[0]["time_s"], 3),
        "entity_type_totals": dict(sorted(all_types.items(), key=lambda x: -x[1])),
    }


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(summaries: list[dict], detail_rows: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[!] openpyxl not installed — skipping Excel output. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2B3A4E")
    alt_fill    = PatternFill("solid", fgColor="EEF2F7")

    def style_header(ws, headers: list[str]) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    def shade_alt_rows(ws, start_row: int = 2) -> None:
        for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
            if i % 2 == 1:
                for cell in row:
                    cell.fill = alt_fill

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    headers1 = [
        "Model",
        "Docs Processed",
        "Total Entities",
        "Avg Entities / Doc",
        "Avg Time / Doc (s)",
        "Avg Time / Doc Steady (s)",
        "Cold Load Time (s)",
        "Min Time (s)",
        "Max Time (s)",
    ]
    style_header(ws1, headers1)
    for s in summaries:
        ws1.append([
            s["model"],
            s["docs_processed"],
            s["total_entities"],
            s["avg_entities_per_doc"],
            s["avg_time_per_doc_s"],
            s["avg_time_per_doc_steady_s"],
            s["cold_load_time_s"],
            s["min_time_s"],
            s["max_time_s"],
        ])
    shade_alt_rows(ws1)
    auto_width(ws1)

    # ── Sheet 2: Per Document ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Per Document")
    headers2 = [
        "Model", "Document", "Pages",
        "Time (s)", "Time / Page (s)",
        "Entities Total", "Includes Model Load",
    ]
    style_header(ws2, headers2)
    for r in detail_rows:
        ws2.append([
            r["model"], r["document"], r["pages"],
            r["time_s"], r["time_per_page_s"],
            r["entities_total"], r["includes_model_load"],
        ])
    shade_alt_rows(ws2)
    auto_width(ws2)

    # ── Sheet 3: Entity Type Breakdown ───────────────────────────────────────
    ws3 = wb.create_sheet("Entity Types")
    # Collect all entity type columns across all models
    all_types = sorted({
        t
        for s in summaries
        for t in s["entity_type_totals"]
    })
    headers3 = ["Model"] + all_types
    style_header(ws3, headers3)
    for s in summaries:
        row = [s["model"]] + [s["entity_type_totals"].get(t, 0) for t in all_types]
        ws3.append(row)
    shade_alt_rows(ws3)
    auto_width(ws3)

    wb.save(out_path)
    print(f"[+] Excel written → {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Benchmark all available NER models against a folder of PDFs."
    )
    parser.add_argument(
        "--docs", required=True, type=Path,
        help="Folder containing PDF files to benchmark against (up to 30 used).",
    )
    parser.add_argument(
        "--out", default="benchmarks/results", type=str,
        help="Output path prefix — .json and .xlsx will be appended (default: benchmarks/results).",
    )
    parser.add_argument(
        "--threshold", default=0.35, type=float,
        help="Presidio confidence threshold (default: 0.35).",
    )
    parser.add_argument(
        "--limit", default=30, type=int,
        help="Max number of PDFs to process (default: 30).",
    )
    args = parser.parse_args()

    models = available_models()
    if not models:
        print("[!] No NER backends installed. Install at least one (spacy, flair, transformers, stanza).")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  NER Benchmark")
    print(f"  Models available : {len(models)}")
    print(f"  Score threshold  : {args.threshold}")
    print(f"{'='*60}\n")

    docs = load_pdfs(args.docs, limit=args.limit)

    all_detail_rows: list[dict] = []
    all_summaries: list[dict] = []

    for model_key in models:
        print(f"\n[→] Model: {model_key}")
        detail_rows = run_model(model_key, docs, score_threshold=args.threshold)
        summary = summarise(model_key, detail_rows)
        all_detail_rows.extend(detail_rows)
        all_summaries.append(summary)

        print(
            f"    Total entities : {summary['total_entities']} | "
            f"Avg/doc : {summary['avg_entities_per_doc']} | "
            f"Avg time (steady) : {summary['avg_time_per_doc_steady_s']}s"
        )

    # ── Write JSON ────────────────────────────────────────────────────────
    out_prefix = Path(args.out)
    out_prefix.parent.mkdir(parents=True, exist_ok=True)

    json_path = out_prefix.with_suffix(".json")
    payload: dict[str, Any] = {
        "config": {
            "score_threshold": args.threshold,
            "docs_folder": str(args.docs),
            "docs_processed": len(docs),
        },
        "summaries": all_summaries,
        "detail": all_detail_rows,
    }
    json_path.write_text(json.dumps(payload, indent=2))
    print(f"\n[+] JSON written  → {json_path}")

    # ── Write Excel ───────────────────────────────────────────────────────
    xlsx_path = out_prefix.with_suffix(".xlsx")
    write_excel(all_summaries, all_detail_rows, xlsx_path)

    print("\n[✓] Benchmark complete.\n")


if __name__ == "__main__":
    main()
