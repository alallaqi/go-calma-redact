#!/usr/bin/env python3
"""Standalone NER benchmark — measures speed and entity detection across all available models.

Completely independent of the main app. Imports gocalma as a library only.

Usage:
    python benchmarks/run_benchmark.py [pdf_folder] [--out results] [--limit 30]

    # Run all available models against the default PDF folder
    python benchmarks/run_benchmark.py /Users/a.a/Downloads/PDFs

    # Run specific models only
    python benchmarks/run_benchmark.py /path/to/pdfs --models "spaCy/en_core_web_lg" "flair/ner-english-large"

Outputs (written to benchmarks/ by default):
    results.json   — full raw data, one entry per model
    results.xlsx   — three sheets: Summary, Per Document, Entity Breakdown
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from collections import defaultdict
from pathlib import Path

# Make gocalma importable when run from the repo root or from benchmarks/.
sys.path.insert(0, str(Path(__file__).parent.parent))

from gocalma.pdf_extract import extract_text
from gocalma.pii_detect import NLP_MODELS, available_models, detect_pii_all_pages


# ---------------------------------------------------------------------------
# PDF loading
# ---------------------------------------------------------------------------

def load_pdfs(folder: Path, limit: int) -> list[tuple[str, bytes]]:
    """Return (filename, bytes) for up to *limit* PDFs in *folder*, sorted by name."""
    pdfs = sorted(folder.glob("*.pdf"))[:limit]
    if not pdfs:
        sys.exit(f"[ERROR] No PDF files found in: {folder}")
    return [(p.name, p.read_bytes()) for p in pdfs]


# ---------------------------------------------------------------------------
# Text extraction (done once, shared across all model runs)
# ---------------------------------------------------------------------------

def extract_all(pdfs: list[tuple[str, bytes]]) -> list[tuple[str, list]]:
    """Pre-extract text from every PDF. Returns (filename, pages) pairs.

    Doing this once upfront means model benchmark timings reflect only NER
    inference, not PDF parsing or OCR.
    """
    results = []
    for name, raw in pdfs:
        print(f"    {name}")
        pages = extract_text(raw)
        results.append((name, pages))
    return results


# ---------------------------------------------------------------------------
# Per-model benchmark
# ---------------------------------------------------------------------------

def benchmark_model(model_key: str, docs: list[tuple[str, list]]) -> dict | None:
    """Run NER on every doc with *model_key*. Returns a structured result dict,
    or None if the model fails to load (e.g. backend not properly installed).

    The first document's time includes model load (lazy initialisation inside
    _get_engine). This is tracked separately so steady-state averages stay
    meaningful.
    """
    doc_results = []

    for i, (name, pages) in enumerate(docs):
        t0 = time.perf_counter()
        try:
            entities = detect_pii_all_pages(pages, model_key=model_key)
        except Exception as exc:
            if i == 0:
                # Failed on first doc — model couldn't load at all; skip it.
                print(f"    [SKIP] {model_key} failed to load: {exc}")
                return None
            # Subsequent docs: record the failure but keep going.
            print(f"    [ERR ] {name}: {exc}")
            doc_results.append({
                "document":        name,
                "pages":           len(pages),
                "time_s":          round(time.perf_counter() - t0, 4),
                "time_per_page_s": 0.0,
                "entity_count":    0,
                "includes_model_load": i == 0,
                "entity_types":    {},
                "error":           str(exc),
            })
            continue

        elapsed = time.perf_counter() - t0

        type_counts: dict[str, int] = defaultdict(int)
        for e in entities:
            type_counts[e.entity_type] += 1

        is_first = i == 0
        note = " (includes model load)" if is_first else ""
        print(f"    [{i + 1:>2}/{len(docs)}] {name}: "
              f"{len(entities)} entities, {elapsed:.2f}s{note}")

        doc_results.append({
            "document":        name,
            "pages":           len(pages),
            "time_s":          round(elapsed, 4),
            "time_per_page_s": round(elapsed / max(len(pages), 1), 4),
            "entity_count":    len(entities),
            "includes_model_load": is_first,
            "entity_types":    dict(type_counts),
        })

    # Summary stats — use all docs for totals, steady-state (docs 2+) for avg timing.
    all_times    = [d["time_s"] for d in doc_results]
    steady_times = all_times[1:] if len(all_times) > 1 else all_times
    all_counts   = [d["entity_count"] for d in doc_results]

    combined_types: dict[str, int] = defaultdict(int)
    for d in doc_results:
        for etype, cnt in d["entity_types"].items():
            combined_types[etype] += cnt

    return {
        "model": model_key,
        "documents": doc_results,
        "summary": {
            "docs_processed":        len(docs),
            "first_doc_time_s":      round(all_times[0], 4),
            "avg_time_per_doc_s":    round(sum(steady_times) / len(steady_times), 4),
            "min_time_s":            round(min(steady_times), 4),
            "max_time_s":            round(max(steady_times), 4),
            "total_entities":        sum(all_counts),
            "avg_entities_per_doc":  round(sum(all_counts) / max(len(all_counts), 1), 1),
            "unique_entity_types":   len(combined_types),
            "entity_type_totals":    dict(
                sorted(combined_types.items(), key=lambda x: -x[1])
            ),
        },
    }


# ---------------------------------------------------------------------------
# JSON output
# ---------------------------------------------------------------------------

def write_json(results: list[dict], path: Path) -> None:
    path.write_text(json.dumps(results, indent=2, ensure_ascii=False))
    print(f"  JSON  → {path}")


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(results: list[dict], path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  Excel → skipped (openpyxl not installed — run: pip install openpyxl)")
        return

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    def style_header(ws) -> None:
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    def autowidth(ws) -> None:
        for col in ws.columns:
            width = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 65)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws.append([
        "Model",
        "Docs Processed",
        "First Doc Time (s)*",
        "Avg Time/Doc (s)",
        "Min Time (s)",
        "Max Time (s)",
        "Total Entities",
        "Avg Entities/Doc",
        "Unique Entity Types",
    ])
    style_header(ws)
    for r in results:
        s = r["summary"]
        ws.append([
            r["model"],
            s["docs_processed"],
            s["first_doc_time_s"],
            s["avg_time_per_doc_s"],
            s["min_time_s"],
            s["max_time_s"],
            s["total_entities"],
            s["avg_entities_per_doc"],
            s["unique_entity_types"],
        ])
    ws.append([])
    ws.append(["* First doc time includes one-off model load. Avg/Min/Max are steady-state (docs 2+)."])
    autowidth(ws)

    # ---- Sheet 2: Per Document -------------------------------------------
    ws2 = wb.create_sheet("Per Document")
    ws2.append([
        "Model",
        "Document",
        "Pages",
        "Total Time (s)",
        "Time per Page (s)",
        "Entities Found",
        "Includes Model Load",
        "Entity Types (type:count)",
    ])
    style_header(ws2)
    for r in results:
        for d in r["documents"]:
            ws2.append([
                r["model"],
                d["document"],
                d["pages"],
                d["time_s"],
                d["time_per_page_s"],
                d["entity_count"],
                "Yes" if d["includes_model_load"] else "No",
                ", ".join(f"{k}:{v}" for k, v in sorted(
                    d["entity_types"].items(), key=lambda x: -x[1]
                )),
            ])
    autowidth(ws2)

    # ---- Sheet 3: Entity Breakdown ---------------------------------------
    ws3 = wb.create_sheet("Entity Breakdown")
    ws3.append(["Model", "Entity Type", "Count", "% of Model Total"])
    style_header(ws3)
    for r in results:
        totals = r["summary"]["entity_type_totals"]
        grand  = r["summary"]["total_entities"] or 1
        for etype, count in totals.items():
            ws3.append([
                r["model"],
                etype,
                count,
                round(count / grand * 100, 1),
            ])
    autowidth(ws3)

    wb.save(path)
    print(f"  Excel → {path}")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Benchmark all available NER models across a folder of PDFs.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "pdf_folder",
        nargs="?",
        default="/Users/a.a/Downloads/PDFs",
        help="Path to folder containing PDF files (default: ~/Downloads/PDFs)",
    )
    parser.add_argument(
        "--out",
        default="benchmarks/results",
        help="Output file prefix without extension (default: benchmarks/results)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=30,
        help="Max number of PDFs to process per model (default: 30)",
    )
    parser.add_argument(
        "--models",
        nargs="*",
        metavar="MODEL_KEY",
        help="Specific model keys to run. Omit to run all available models.",
    )
    args = parser.parse_args()

    pdf_folder = Path(args.pdf_folder)
    if not pdf_folder.is_dir():
        sys.exit(f"[ERROR] Not a directory: {pdf_folder}")

    # Resolve models to benchmark
    avail = available_models()
    if not avail:
        sys.exit("[ERROR] No NER backends are installed. Check requirements.txt.")

    models_to_run = avail
    if args.models:
        models_to_run = {k: v for k, v in avail.items() if k in args.models}
        missing = [m for m in args.models if m not in avail]
        if missing:
            print(f"[WARN] Not available (skipped): {missing}")
        if not models_to_run:
            sys.exit("[ERROR] None of the requested models are available.")

    # Header
    print("\n" + "=" * 60)
    print("  NER MODEL BENCHMARK")
    print("=" * 60)
    print(f"  PDF folder : {pdf_folder}")
    print(f"  PDF limit  : {args.limit}")
    print(f"  Models     : {len(models_to_run)}")
    for k in models_to_run:
        print(f"    - {k}")
    print()

    # Step 1: load PDFs
    print(f"[1/3] Loading PDFs...")
    pdfs = load_pdfs(pdf_folder, args.limit)
    print(f"  Found {len(pdfs)} PDF(s)\n")

    # Step 2: extract text once
    print(f"[2/3] Pre-extracting text (done once, shared across all models)...")
    docs = extract_all(pdfs)
    total_pages = sum(len(p) for _, p in docs)
    print(f"  {total_pages} pages across {len(docs)} document(s)\n")

    # Step 3: benchmark
    print(f"[3/3] Running benchmarks...\n")
    all_results: list[dict] = []
    skipped: list[str] = []
    for i, (model_key, _) in enumerate(models_to_run.items()):
        print(f"  [{i + 1}/{len(models_to_run)}] {model_key}")
        result = benchmark_model(model_key, docs)
        if result is None:
            skipped.append(model_key)
        else:
            all_results.append(result)
        print()

    # Write outputs
    out_prefix = Path(args.out)
    out_prefix.parent.mkdir(parents=True, exist_ok=True)
    print("[OUTPUT]")
    write_json(all_results, out_prefix.with_suffix(".json"))
    write_excel(all_results, out_prefix.with_suffix(".xlsx"))

    # Quick summary table
    print()
    print("=" * 85)
    print(f"  {'Model':<44} {'Avg/Doc':>9} {'Min':>7} {'Max':>7} {'Entities':>9} {'Types':>6}")
    print("-" * 85)
    for r in all_results:
        s = r["summary"]
        print(
            f"  {r['model']:<44} "
            f"{s['avg_time_per_doc_s']:>8.2f}s "
            f"{s['min_time_s']:>6.2f}s "
            f"{s['max_time_s']:>6.2f}s "
            f"{s['total_entities']:>9} "
            f"{s['unique_entity_types']:>6}"
        )
    print("=" * 85)
    print("  Note: Avg/Min/Max exclude the first doc (which includes model load time).")
    print("  See First Doc Time column in results.xlsx for cold-start timing.")
    if skipped:
        print(f"\n  Skipped (failed to load): {', '.join(skipped)}")
    print()


if __name__ == "__main__":
    main()
